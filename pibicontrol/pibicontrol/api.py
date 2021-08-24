# -*- coding: utf-8 -*-
# Copyright (c) 2020, PibiCo and contributors
# For license information, please see license.txt
from __future__ import unicode_literals
import frappe
import datetime, json, subprocess
from frappe import _, msgprint

from frappe.utils import cstr, time_diff_in_seconds, get_files_path
from frappe.core.doctype.sms_settings.sms_settings import send_sms
from pibicontrol.pibicontrol.doctype.telegram_settings.telegram_settings import send_telegram
from pibicontrol.pibicontrol.doctype.mqtt_settings.mqtt_settings import send_mqtt

import paho.mqtt.client as mqtt
import os, ssl, urllib, time, json
from frappe.utils.password import get_decrypted_password

DATE_FORMAT = "%Y-%m-%d"
TIME_FORMAT = "%H:%M:%S.%f"
CHART_FORMAT = "%H:%M"
DATETIME_FORMAT = DATE_FORMAT + " " + TIME_FORMAT

@frappe.whitelist()
def get_image(slide):
  """ Get from database all images associated  """
  data = frappe.db.sql("""
		SELECT * FROM `tabWebsite Slideshow Item` WHERE  parent=%s and docstatus<2""", slide, True)
  return data

@frappe.whitelist()
def switch(value, action):
  sensor = frappe.get_doc("Sensor", value)
  trigger = sensor.trigger_pin
  info = sensor.info_pin
  pos = value.find("-")
  topic = []
  topic.append(sensor.hostname + "/mqtt")
  type = value[:pos]
  msg = type + "_" + action + "(" + str(trigger) + "," + str(info) + ")"
  send_mqtt(topic, cstr(msg))

@frappe.whitelist()
def mqtt_command(host, action):
  topic = []
  topic.append(host + "/mqtt")
  send_mqtt(topic, cstr(action))
  
@frappe.whitelist()
def get_chart_dataset (doc):
  label = []
  main_read = []
  second_read = []
  third_read = []
  second_var = ''
  second_uom = ''
  third_var = ''
  third_uom = ''
  data = frappe.get_doc("Sensor Log", doc)
  #print("data = {}".format(frappe.as_json(data)))
  if data.log_item:
    if len(data.log_item) > 0:
      variable = data.log_item[0].main_reading
      uom = data.log_item[0].uom
      if "cpu-" in data.sensor:
        second_var = "mem_pct"
        second_uom = "%"
        third_var = "disk_pct"
        third_uom = "%"       
      elif "th-" in data.sensor:
        second_var = "env_humid"
        second_uom = "%"
      elif "ut-" in data.sensor:
        second_var = "env_vol"
        second_uom = "l"
      elif "soil-" in data.sensor:
        second_var = "soil_temp"
        second_uom = "C"
      for item in data.log_item:
        main_read.append(item.value)
        label.append(item.datadate.strftime(CHART_FORMAT))
        payload = json.loads(item.payload)
        if "cpu-" in data.sensor:
          second_read.append(payload['payload']['mem']['mem_pct']) 
          third_read.append(payload['payload']['disk']['disk_pct'])
        elif "th-" in data.sensor:
          second_read.append(payload['payload']['reading']['val_humid'])  
        elif "ut-" in data.sensor:
          second_read.append(payload['payload']['reading']['val_vol'])       
        elif "soil-" in data.sensor:
          second_read.append(payload['payload']['reading']['val_temp'])            
    
  return {
    'label': label,
    'main_read': main_read,
    'variable': variable,
    'uom': uom,
    'second_read': second_read,
    'second_var': second_var,
    'second_uom': second_uom,
    'third_read': third_read,
    'third_var': third_var,
    'third_uom': third_uom
  }

def check_off_schedule(sensor):
  off_time = datetime.datetime.now().time().strftime(TIME_FORMAT)
  from_time = 0
  to_time = 0
  ## Calculate difference in seconds from current time to Off Schedule From Time
  if sensor.off_from_time:
    from_time = time_diff_in_seconds(off_time, str(sensor.off_from_time))
  ## Calculate difference in seconds from current time to Off Schedule To Time  
  if sensor.off_to_time:
    to_time = time_diff_in_seconds(off_time, str(sensor.off_to_time))
  ## if off_time is between from time and to time not sending messages  
  if from_time > 0 and to_time < 0:
    return True
  else:
    return False      

def get_alert(variable, name):
  start = True
  ## Get active alerts for the Sensor based on not having to_time value in record
  last_log = frappe.get_list(
    doctype = "Alert Log",
    fields = ['*'],
    filters = [['date', '=', datetime.datetime.now().strftime(DATE_FORMAT)],['docstatus', '<', 2], ['sensor', '=', name], ['variable','=', variable]],
    order_by = 'date desc',
    limit_start = 0,
    limit_page_length = 1
  )
  ## Get last Alert Log
  active_alert = None
  if last_log:
    if len(last_log) > 0:
      active_alert = frappe.get_doc("Alert Log", last_log[0].name)
      ## If recorded alert is not closed then will be a starting alert
      if active_alert:
        if not active_alert.alert_item[len(active_alert.alert_item)-1].to_time:
          start = False
    ## Script (future) for not having log
      
  return (active_alert, start)

def reschedule_alert():
  ## Get active alerts not having to_time value in record
  alert_list = frappe.get_list(
    doctype = "Alert Item",
    fields = ['*'],
    filters = [['parent', 'like', frappe.utils.add_days(frappe.utils.getdate(), -1).strftime("%y%m%d") + "_%"],['docstatus', '<', 2]]
  )
  for alert in alert_list:
    if not alert.to_time:
      alert_log = frappe.get_doc("Alert Log", alert.parent)
      for row in alert_log.alert_item:
        if not row.to_time:
          ## Write Data for new Alert
          alert_json = row
          ## Code for creating new doc
          alert_new = frappe.new_doc('Alert Log')
          alert_new.sensor = alert_log.sensor
          alert_new.date = frappe.utils.getdate().strftime(DATE_FORMAT)
          alert_new.variable = alert_log.variable
          ## Adds log to array  
          alert_new.append("alert_item", alert_json)
          alert_list = frappe.get_list(
            doctype = "Alert Log",
            fields = ['name'],
            filters = [['sensor', '=', alert_log.sensor], ['docstatus', '<', 2], ['date', '=', frappe.utils.getdate().strftime(DATE_FORMAT)]]
          )
          if not alert_list:
            alert_new.save()
            print(_("[INFO] Rescheduled Old Alert"))
  
def mng_alert(sensor, variable, value, start, alert_log):
  """ doc is Sensor, variable is variable, value is actual value for variable, start is True or False whether the alert is beginning or ending, active_alert is active_alert """
  if not sensor.disabled and sensor.alerts_active:
    ## Prepare message to send
    if start:
      if variable == "last_seen":
        msg = sensor.name + " offline. Please check!"
      elif variable == "offline":
        msg = sensor.name + " web service " + variable + ". Please check!"
      elif variable == "status":
        msg = sensor.name + ". Sensor has been switched on. Please check!"
      elif variable == "pir_alarm":
        msg = sensor.name + ". Sensor detected presence. Please check!"
      else:
        msg = sensor.name + " detected abnormal value " + str(value) + " in " + variable + ". Please check!"
    else:
      if variable == "last_seen":
        msg = sensor.name + " again online. Rest easy!"
      elif variable == "offline":
        msg = sensor.name + " web services again online. Rest easy!"
      elif variable == "status":
        msg = sensor.name + ". Sensor has been switched off. Rest easy!"
      elif variable == "pir_alarm":
        msg = sensor.name + ". Normality restored. Rest easy!"
      else:
        msg = sensor.name + " recovered normal value " + str(value) + " in " + variable + ". Rest easy!"
    ## Check Active Channels and prepare a thread to alert
    ## Fill main parameters
    doSend = True
    datadate = datetime.datetime.now()
    by_sms = by_telegram = by_mqtt = by_email = 0
    sms_list = []
    telegram_list = []
    mqtt_list = []
    email_list = []
    if sensor.sms_alert and sensor.sms_recipients != '':
      alert_sms = "SMS from " + msg 
      by_sms = 1
      sms_list = sensor.sms_recipients.split(",")
    if sensor.telegram_alert and sensor.telegram_recipients != '':
      alert_telegram = "Telegram from " + msg
      by_telegram = 1
      telegram_list = sensor.telegram_recipients.split(",")
    if sensor.mqtt_alert and sensor.mqtt_recipients != '':
      if sensor.mqtt_command:
        alert_mqtt = sensor.mqtt_command
      else:  
        alert_mqtt = "MQTT from " + msg
      by_mqtt = 1
      mqtt_list = sensor.mqtt_recipients.split(",")
    if sensor.email_alert and sensor.email_recipients != '':
      if start:
        subject = "Alert from " + sensor.name
      else:
        subject = "Finished Alert from " + sensor.name
      alert_email = "From " + msg
      by_email = 1
      email_list = sensor.email_recipients.split(",")
    ## Write Data for new Alert
    alert_json = {
      "variable": variable,
      "value": value,
      "from_time": datadate.strftime(DATETIME_FORMAT),
      "to_time": None,
      "by_sms": by_sms,
      "by_telegram": by_telegram,
      "by_mqtt": by_mqtt,
      "by_email": by_email
    }
    if alert_log:
      ##if len(alert_log) > 0:
      last_alert = alert_log.alert_item[len(alert_log.alert_item)-1]
      if alert_log.date.strftime(DATE_FORMAT) == datadate.strftime(DATE_FORMAT):
        ## Code for update child item in doc
        if not last_alert.to_time and not start:
          last_alert.to_time = datadate.strftime(DATETIME_FORMAT)
          alert_log.save()
          frappe.msgprint(_("[INFO] Alert Finished"))
        elif last_alert.to_time and start:
          alert_log.append("alert_item", alert_json)
          alert_log.save()
          frappe.msgprint(_("[INFO] New Alert open"))
        elif not last_alert.to_time and start:
          doSend = False
          frappe.msgprint(_("[INFO] Alert already open"))
    else:
      if start:
        ## Code for creating new doc
        alert_log = frappe.new_doc('Alert Log')
        alert_log.sensor = sensor.name
        alert_log.date = datadate.strftime(DATETIME_FORMAT)
        alert_log.variable = variable
        ## Adds log to array  
        alert_log.append("alert_item", alert_json)
        alert_log.save()
        frappe.msgprint(_("[INFO] Created New Log"))
    ## Sending messages to recipients provided that alerts are not on off schedule
    if check_off_schedule(sensor):
      doSend = False
    ## Send messages
    if len(sms_list) > 0 and doSend:
      send_sms(sms_list, cstr(alert_sms))
    if len(telegram_list) > 0 and doSend:
      send_telegram(telegram_list, cstr(alert_telegram))
    if len(mqtt_list) > 0 and doSend:
      send_mqtt(mqtt_list, cstr(alert_mqtt))
    if len(email_list) > 0 and doSend:
      frappe.sendmail(recipients=email_list, subject=subject, message=cstr(alert_email))

def ping_devices_via_mqtt():
  sensors = frappe.get_list(
    doctype = "Sensor",
    fields = ['*'],
    filters = [['docstatus', '<', 2], ['disabled', '=', 0]]
  )
  if sensors:
    command = "is_alive"
    strboot = "boot"
    for device in sensors:
      hostname = []
      hostname.append(device.hostname + "/mqtt")
      lastseen = device.lastseen
      now = datetime.datetime.now()
      pos = device.sensor_shortcut.find("-")
      strcmd = "start_" + device.sensor_shortcut[:pos]
      cmd = "take_" + device.sensor_shortcut[:pos]
      
      if lastseen:
        time_minutes = (now - lastseen).total_seconds()/60
      else:
        time_minutes = 12
        
      if time_minutes <= 5:
        ## Check active last_seen alerts to close
        (active_alert, start) = get_alert("last_seen", device.name)
        if start == False:
          mng_alert(device, 'last_seen', 1, start, active_alert)
      elif 15 >= time_minutes > 5:
        ## Order through MQTT to restart supervisor daemons
        if "cpu-" in device.sensor_shortcut:
          send_mqtt(hostname, cstr(command))
        else:
          send_mqtt(hostname, cstr(cmd))
      elif time_minutes > 15:
        ## Not receiving data from device
        ## Send alert last_seen 0 from_time
        send_mqtt(hostname, cstr(strcmd))
        (active_alert, start) = get_alert("last_seen", device.name)
        if start == True:
          mng_alert(device, 'last_seen', 0, start, active_alert)
      elif time_minutes > 30:
        send_mqtt(hostname, cstr(strboot))

def create_xls_report():
  ## Import needed libraries
  ## If not install pathlib sudo apt-get install python-pathlib or sudo pip3 install pathlib
  from pathlib import Path
  ## If not install openpyxl sudo pip3 install openpyxl
  import openpyxl
  from openpyxl.styles import Font
  from openpyxl.styles.colors import Color
  import shutil
  ## If not installed sudo pip3 install tzlocal
  from tzlocal import get_localzone
  from time import gmtime, strftime
  import urllib3
  ## if not install sudo pip3 install pyocclient
  import owncloud
  
  ## Calculate actual system date
  otoday = datetime.datetime.today()
  ## We not use day, month or year datetime functions for keeping variables as string
  oyear = otoday.strftime("%Y")
  omonth = otoday.strftime("%m")
  oday = int(otoday.strftime("%d"))
  ohour = int(otoday.strftime("%H"))
  ominutes = int(otoday.strftime("%M"))
  
  ## Get yesterday's date
  DAY = datetime.timedelta(1)
  local_tz = get_localzone()   # get local timezone
  now = datetime.datetime.now(local_tz) # get timezone-aware datetime object
  day_ago = local_tz.normalize(now - DAY) # exactly 24 hours ago, time may differ
  naive = now.replace(tzinfo=None) - DAY # same time
  yesterday = local_tz.localize(naive, is_dst=None) # but elapsed hours may differ
  yyear = yesterday.strftime("%Y")
  ymonth = yesterday.strftime("%m")
  yday = int(yesterday.strftime("%d"))
  ## Set yesterday and actual day
  daybefore = yyear + '-' + ymonth + '-' + yesterday.strftime("%d")
  current = oyear + '-' + omonth + '-' + otoday.strftime("%d")
  #Monthly report template
  dir_src = ("/home/erpnext/erpnext-prd/sites/" + frappe.get_site_path().replace("./","") + "/private/files/")
  report_template = frappe.get_doc(
    doctype = 'File',
    file_name = 'Pharmacy_Report_v0.xlsx')
  template_file = report_template.file_name
  src_file = dir_src + template_file
  ## Get all sensors from group PibiFarma, not cpus and grouped by client
  pharma_sensors = frappe.db.sql("""
    SELECT * 
    FROM `tabSensor` 
    WHERE sensor_group=%s AND docstatus<2 AND NOT sensor_shortcut LIKE %s AND NOT disabled
  """,("PibiFarma","cpu-%"), True)
  ## Get unique values for clients
  clients = set()
  for val in pharma_sensors:
    clients.add(val.client)
  clients = list(clients)
  for user in clients:
    client = frappe.get_doc("Client", user)
    organization = client.organization
    if not organization:
      organization = ''
    ## Check whether monthly report exists and create
    if ymonth != omonth:
      omonth = ymonth
    if yyear != oyear:
      oyear = yyear
    ofile = str(oyear) + str(omonth) + '_farmacia_' + client.user + '.xlsx'
    dst_file = dir_src + "reports/" + ofile
    oreport = Path(dst_file)
    if oreport.exists():
      ## Filepath exists
      isFile = True
    else:
      ## File create
      shutil.copy(src_file, dst_file)
    ## Loads existing or recently created monthly workbook 
    wbook = openpyxl.load_workbook(dst_file)
    sheet = wbook['Farmacia']
    ## Writes Organization Name on header
    sheet.cell(row= 2, column=4).value = str(organization)
    sheet.cell(row= 2, column= 54).value = str(omonth)
    sheet.cell(row= 2, column= 62).value = str(oyear)
    ## Select Fridges and Zones from Sensors property of Client
    selectFridge = []
    selectZone = []
    for sensor in pharma_sensors:
      if "temp-" in sensor.sensor_shortcut and client.user == sensor.client:
        selectFridge.append(sensor)
      if "th-" in sensor.sensor_shortcut and client.user == sensor.client:
        selectZone.append(sensor)
    ## Write Values for only 1 fridge
    for nFridges, fridge in enumerate(selectFridge, start=1): 
      if nFridges == 1:
        ## Select Sensor Log for last day provided that script is running at 0:00 current day
        sensor_log = frappe.db.sql("""
          SELECT *
          FROM `tabSensor Log`
          WHERE sensor=%s AND docstatus<2 AND date=%s
          LIMIT 1
        """, (fridge.sensor_shortcut, daybefore), True)
        ## Select Item Log for Sensor on Yesterday
        log = frappe.db.sql("""
          SELECT *
          FROM `tabLog Item`
          WHERE parent=%s
        """, sensor_log[0].name, True)
        maximo = sensor_log[0]['max']
        minimo = sensor_log[0]['min']
        count = sensor_log[0].points
        location = fridge.room
        fridgeserial = fridge.serial   
        ## Select last data from a range in morning and another one for the afternoon
        for row, item in enumerate(log, start=0):
          if datetime.datetime.strptime(daybefore + " 06:00:00.00", '%Y-%m-%d %H:%M:%S.%f') < item.datadate < datetime.datetime.strptime(daybefore + " 12:00:00.00", '%Y-%m-%d %H:%M:%S.%f'):
            datamorning = item
          elif datetime.datetime.strptime(daybefore + " 17:00:00.00", '%Y-%m-%d %H:%M:%S.%f') < item.datadate < datetime.datetime.strptime(daybefore + " 23:00:00.00", '%Y-%m-%d %H:%M:%S.%f'):
            datafternoon = item    
        ## Write data to Sheet
        #print("M:" + str(datamorning.value) + " T:" + str(datafternoon.value))
        sheet.cell(row= 6, column= 2).value = location + " (" + str(fridgeserial) + ")"
        #morning data
        if datamorning and -2 < float(datamorning.value) < 12:
          sheet.cell(row = 19 - round(float(datamorning.value),0) , column = yday*2+3).value = "x"
        #afternoon data
        if datafternoon and -2 < float(datafternoon.value) < 12:
          sheet.cell(row = 19 - round(float(datafternoon.value),0) , column = yday*2+4).value = "y"
        #maximum reading
        if maximo != -float('inf'):
          sheet.cell(row = 22 , column = yday*2+3).value = str(int(round(maximo,0)))
        if float(maximo) > 8:
          sheet.cell(row = 22 , column = yday*2+3).font = Font(color = "FF0000")
        #minimum reading
        if minimo != float('inf'):
          sheet.cell(row = 23 , column = yday*2+3).value = str(int(round(minimo,0)))
        if float(minimo)<2:
          sheet.cell(row = 23 , column = yday*2+3).font = Font(color = "FF0000")
        #print("fridge from " + organization + " (" + fridgeserial + ") Max:" + str(maximo) + " Min:" + str(minimo) + " count(" + str(count) + ") on " + daybefore)
    ## Write Values for zones
    for nZones, zone in enumerate(selectZone, start=1):     
      ## Select Sensor Log for last day provided that script is running at 0:00 current day
      sensor_log = frappe.db.sql("""
        SELECT *
        FROM `tabSensor Log`
        WHERE sensor=%s AND docstatus<2 AND date=%s
        LIMIT 1
      """, (zone.sensor_shortcut, daybefore), True)
      ## Select Item Log for Sensor on Yesterday
      log = frappe.db.sql("""
        SELECT *
        FROM `tabLog Item`
        WHERE parent=%s
      """, sensor_log[0].name, True)
      maximo = sensor_log[0]['max']
      minimo = sensor_log[0]['min']
      count = sensor_log[0].points
      locarea = zone.room
      zoneserial = zone.serial
      ## Write header for zone sensor
      sheet.cell(row= 24 + (nZones - 1)*4, column= 2).value = locarea + " (" + str(zoneserial) + ")"
      ## Calculate humidity values from payload
      counter = 0
      avgval = 0
      humidity = 0
      for item in log:
        counter += 1
        valitem = json.loads(item.payload)
        valhumid = (valitem['payload']['reading']['val_humid'])
        avgval = avgval + valhumid
      if counter > 0:
        humidity = round(avgval/counter,1)  
      ## Write values to cells
      ## maximum reading
      if maximo != -float('inf'):
        sheet.cell(row = 25 + (nZones - 1)*4, column = yday*2 + 3).value = str(int(round(maximo,0)))
      if float(maximo)>25:
        sheet.cell(row = 25 + (nZones - 1)*4, column = yday*2 + 3).font = Font(color = "FF0000")
      ## minimum reading
      if minimo != float('inf'):
        sheet.cell(row = 26 + (nZones - 1)*4 , column = yday*2 + 3).value = str(int(round(minimo,0)))
      if float(minimo)<12:
        sheet.cell(row = 26 + (nZones - 1)*4, column = yday*2 + 3).font = Font(color = "FF0000")
      ## humidity reading
      sheet.cell(row = 27 + (nZones - 1)*4 , column = yday*2 + 3).value = str(int(round(humidity,0)))
      if float(humidity) > 65:
        sheet.cell(row = 27 + (nZones - 1)*4, column = yday*2 + 3).font = Font(color = "FF0000")
      #print("zone from " + organization + " (" + zoneserial + ") Max:" + str(maximo) + " Min:" + str(minimo) + " H:" + str(humidity) + " count(" + str(counter) + ") on " + daybefore)
            
    ## Close workbook and upload to cloud
    wbook.save(dst_file)
    ## Once saved the workbook upload to nextcloud
    ## Better substitute by recommendations in https://urllib3.readthedocs.io/en/latest/advanced-usage.html     ## ssl-warnings
    urllib3.disable_warnings()
    path = ('/Informes/' + oyear + '/' + ofile)
    remotepath = path.encode("ascii", "ignore").decode("ascii").strip()
    nc_server = client.cloud_gateway
    nc_user = client.cloud_user
    nc_secret = get_decrypted_password("Client", client.user, "cloud_secret", False)
    nextcloud = owncloud.Client(nc_server)
    nextcloud.login(nc_user, nc_secret)
    nextcloud.put_file(remotepath, dst_file)

def check_web_services():
  ## In order to remove error on self-signed-certificates or overpassed
  ssl._create_default_https_context = ssl._create_unverified_context
  ## Pinging Function
  def pinger_urllib(host):
    try:
      start_time = time.time()
      urllib.request.urlopen(host).read()
      return (time.time() - start_time) * 1000.0
    except:
      return float('inf')
  ## Task Main Function    
  def task(m, device):
    delay = float(pinger_urllib(m))
    if delay == float('inf'):
      ## Check if offline alert is active and send alert and detailed email
      (active_alert, start) = get_alert("offline", device.name)
      if start == True:
        mng_alert(device, 'offline', 1, start, active_alert)
        #print(m + ' from ' + device.sensor_shortcut + ' is offline')
        email_list = []
        alert_email = "Web Service " + m + " from " + device.sensor_shortcut + " is offline. Please check!"
        subject = m + " Offline" 
        email_list.append(device.client)
        ## Sending messages to recipients provided that alerts are not on off schedule
        if not check_off_schedule(device):
          frappe.sendmail(recipients=email_list, subject=subject, message=cstr(alert_email))
    else:
      ## Check if offline alert is active and close
      (active_alert, start) = get_alert("offline", device.name)
      if start == False:
        mng_alert(device, 'offline', 1, start, active_alert)
        #print(m + ' from ' + device.sensor_shortcut + ' again online')
        email_list = []
        alert_email = "Web Service " + m + " from " + device.sensor_shortcut + " is online again. Rest Easy!"
        subject = m + " Online Again" 
        email_list.append(device.client)
        if not check_off_schedule(device):
          frappe.sendmail(recipients=email_list, subject=subject, message=cstr(alert_email))
    print('%-30s %-15s %5.0f [ms]' % (m, device.sensor_shortcut, delay))
  ## Get all web services from active sensors
  sensors = frappe.db.sql("""
    SELECT *
    FROM `tabSensor`
    WHERE disabled=%s AND docstatus<2 AND NOT web_services=%s AND alerts_active=%s 
    """, (0, '', 1), True)
  ## Ping every sensor for its web services in list  
  for sensor in sensors:
    if "," in sensor.web_services:
      sensor_list = list(sensor.web_services.split(","))
      for item in sensor_list:
        task(item, sensor)
    else:
      task(sensor.web_services, sensor)
    time.sleep(0.15)  